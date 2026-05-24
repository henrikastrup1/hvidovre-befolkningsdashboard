import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook('D:/hermes-usb-portable-main/data/skills/Danmarks Statistik API/hvidovre_folk1am.xlsx')
ws3 = wb['Opdatering']
for row in ws3.iter_rows():
    for cell in row:
        cell.value = None

ws3.column_dimensions['A'].width = 55
ws3.column_dimensions['B'].width = 85

info = [
    ('', ''),
    ('SADAN TILFOJER DU EN OPDATERINGS-KNAP', ''),
    ('', ''),
    ('Trin 1:', 'Gem arket som .xlsm (Excel-makroaktiveret projektmappe)'),
    ('Trin 2:', 'Aaben Visual Basic Editor (Alt+F11)'),
    ('Trin 3:', 'File > Import File > vaalg DST_Refresh.bas (samme mappe)'),
    ('Trin 4:', 'Luk VBA editoren. Ga til: Udvikler > Indsaet > Knap (Formularstyring)'),
    ('Trin 5:', 'Tegn en knap i arket -- vaalg RefreshDSTData som makro'),
    ('Trin 6:', 'Hojreklik knappen > Rediger tekst > skriv f.eks. "Opdater data fra DST"'),
    ('Trin 7:', 'Fremover: klik paa knappen = data hentes fra Danmarks Statistik!'),
    ('', ''),
    ('ALTERNATIV: Opdater uden makro (Power Query)', ''),
    ('', ''),
    ('Data > Hent data > Fra internettet > indsaet URL nedenfor', ''),
    ('Filtrer "Alder i alt" fra > Indlaes', ''),
    ('Fremtidig opdatering: Ctrl+Alt+F5', ''),
    ('', ''),
    ('API-URL:', 'https://api.statbank.dk/v1/data/FOLK1AM/CSV?OMR%C3%85DE=167&K%C3%98N=TOT&ALDER=*&TID=*&lang=da'),
    ('', ''),
    ('API detaljer', ''),
    ('Endpoint:', 'api.statbank.dk/v1/data/{tabel}/{format}'),
    ('Tabel:', 'FOLK1AM -- Befolkning pr. maaned'),
    ('Kommune:', 'OMRADE=167 (Hvidovre)'),
    ('Kon:', 'KON=TOT'),
    ('Aldre:', 'ALDER=* (alle aldre)'),
    ('Tid:', 'TID=* (alle perioder)'),
    ('Format:', 'CSV -- i URL-stien (IKKE som query-parameter)'),
]

for i, (lbl, val) in enumerate(info):
    r = 3 + i
    if lbl and not val:
        ws3.cell(row=r, column=1, value=lbl).font = Font(name='DM Sans', size=11, bold=True, color='1A4D2E')
    elif lbl.startswith('Trin'):
        ws3.cell(row=r, column=1, value=lbl).font = Font(name='DM Mono', size=10, bold=True, color='4A4A44')
    elif lbl in ('API-URL:', 'Endpoint:', 'Tabel:', 'Kommune:', 'Kon:', 'Aldre:', 'Tid:', 'Format:'):
        ws3.cell(row=r, column=1, value=lbl).font = Font(name='DM Mono', size=10, bold=True, color='4A4A44')
    elif lbl:
        ws3.cell(row=r, column=1, value=lbl).font = Font(name='DM Sans', size=10, bold=True, color='4A4A44')
    if val:
        ws3.cell(row=r, column=2, value=val).font = Font(name='DM Sans', size=10, color='1C1C1A')

out = 'D:/hermes-usb-portable-main/data/skills/Danmarks Statistik API/hvidovre_folk1am.xlsm'
wb.save(out)
print(f'Saved: {out}')
