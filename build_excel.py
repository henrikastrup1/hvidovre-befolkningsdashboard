import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv, io

csv_path = "D:/hermes-usb-portable-main/data/skills/Danmarks Statistik API/folk1am_raw.csv"
with open(csv_path, encoding='utf-8-sig') as f:
    csv_text = f.read()

rows = list(csv.DictReader(io.StringIO(csv_text), delimiter=';'))
# Filter out "Alder i alt" rows
rows = [r for r in rows if 'alt' not in r['ALDER'].lower()]
periods = sorted(set(r['TID'] for r in rows))
ages = sorted(set(r['ALDER'] for r in rows), key=lambda a: int(a.split()[0]))
data = {(r['ALDER'], r['TID']): int(r['INDHOLD']) for r in rows}

wb = openpyxl.Workbook()
GREEN = '1A4D2E'; GREEN_MID = '2E7D54'; WHITE = 'FFFFFF'
BORDER = 'E2DDD4'; TEXT = '1C1C1A'; MID = '4A4A44'; MUTED = '8A8A7E'
hdr_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
hdr_font = Font(name='DM Sans', size=10, bold=True, color=WHITE)
data_font = Font(name='DM Sans', size=10, color=TEXT)
mono_font = Font(name='DM Mono', size=10, color=TEXT)
green_fill = PatternFill(start_color=GREEN_MID, end_color=GREEN_MID, fill_type='solid')
alt_fill = PatternFill(start_color='F3F0EB', end_color='F3F0EB', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color=BORDER), right=Side(style='thin', color=BORDER),
    top=Side(style='thin', color=BORDER), bottom=Side(style='thin', color=BORDER)
)

# ===== SHEET 1: Data =====
ws1 = wb.active; ws1.title = 'Data'
ws1.merge_cells(f'A1:{get_column_letter(len(periods)+3)}1')
ws1['A1'] = 'Hvidovre Kommune · Befolkning pr. alder · Danmarks Statistik FOLK1AM'
ws1['A1'].font = Font(name='DM Serif Display', size=16, color=GREEN)
ws1.row_dimensions[1].height = 30
ws1.merge_cells(f'A2:{get_column_letter(len(periods)+3)}2')
ws1['A2'] = f'Data: {periods[0]}–{periods[-1]} ({len(periods)} mdr) · Seneste: {periods[-1]} · Opdater: Ctrl+Alt+F5'
ws1['A2'].font = Font(name='DM Mono', size=9, color=MUTED)

ph = [f"{p[:4]}M{int(p[5:]):02d}" for p in periods]
headers = ['Alder'] + ph + [f'{periods[-1][:4]}M{int(periods[-1][5:]):02d}–{periods[0][:4]}M{int(periods[0][5:]):02d}', 'Ændring %']
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=4, column=col, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin_border

for i, age in enumerate(ages):
    row = 5 + i; an = int(age.split()[0])
    ws1.cell(row=row, column=1, value=an).font = mono_font
    ws1.cell(row=row, column=1).alignment = Alignment(horizontal='center'); ws1.cell(row=row, column=1).border = thin_border
    for j, p in enumerate(periods):
        v = data.get((age, p), 0)
        c = ws1.cell(row=row, column=2+j, value=v); c.font = data_font; c.alignment = Alignment(horizontal='right'); c.border = thin_border; c.number_format = '#,##0'
    jv = data.get((age, periods[0]), 0); av = data.get((age, periods[-1]), 0); ch = av - jv; cp = ch / jv if jv else 0
    c = ws1.cell(row=row, column=len(periods)+2, value=ch)
    c.font = Font(name='DM Sans', size=10, color='C0392B' if ch < 0 else GREEN_MID, bold=True)
    c.alignment = Alignment(horizontal='right'); c.border = thin_border; c.number_format = '+#,##0;-#,##0;0'
    c = ws1.cell(row=row, column=len(periods)+3, value=cp)
    c.font = Font(name='DM Sans', size=10, color='C0392B' if ch < 0 else GREEN_MID, bold=True)
    c.alignment = Alignment(horizontal='right'); c.border = thin_border; c.number_format = '+0.0%;-0.0%;0.0%'
    if i % 2 == 1:
        for col in range(1, len(headers)+1): ws1.cell(row=row, column=col).fill = alt_fill
    ws1.row_dimensions[row].height = 18

# Total row
tr = 5 + len(ages)
ws1.cell(row=tr, column=1, value='I alt').font = Font(name='DM Sans', size=10, bold=True, color=WHITE)
ws1.cell(row=tr, column=1).fill = green_fill; ws1.cell(row=tr, column=1).alignment = Alignment(horizontal='center'); ws1.cell(row=tr, column=1).border = thin_border
for j, p in enumerate(periods):
    v = sum(data.get((a, p), 0) for a in ages)
    c = ws1.cell(row=tr, column=2+j, value=v); c.font = Font(name='DM Sans', size=10, bold=True, color=WHITE); c.fill = green_fill; c.alignment = Alignment(horizontal='right'); c.border = thin_border; c.number_format = '#,##0'
tj = sum(data.get((a, periods[0]), 0) for a in ages)
ta = sum(data.get((a, periods[-1]), 0) for a in ages)
tc = ta - tj; tp = tc / tj if tj else 0
for col, v, fmt in [(len(periods)+2, tc, '+#,##0;-#,##0;0'), (len(periods)+3, tp, '+0.0%;-0.0%;0.0%')]:
    c = ws1.cell(row=tr, column=col, value=v); c.font = Font(name='DM Sans', size=10, bold=True, color=WHITE); c.fill = green_fill; c.alignment = Alignment(horizontal='right'); c.border = thin_border; c.number_format = fmt

ws1.column_dimensions['A'].width = 8
for j in range(len(periods)): ws1.column_dimensions[get_column_letter(2+j)].width = 9
ws1.column_dimensions[get_column_letter(len(periods)+2)].width = 20
ws1.column_dimensions[get_column_letter(len(periods)+3)].width = 12
ws1.freeze_panes = 'B5'

# ===== SHEET 2: Aldersgrupper =====
ws2 = wb.create_sheet('Aldersgrupper')
ws2.merge_cells('A1:F1'); ws2['A1'] = 'Budgetrelevante aldersgrupper'
ws2['A1'].font = Font(name='DM Serif Display', size=14, color=GREEN)
bg = [('0-2 år (vuggestue)', 0, 2, 'Dagtilbud — vuggestuepladser'), ('3-5 år (børnehave)', 3, 5, 'Dagtilbud — børnehavepladser'), ('6-15 år (skole)', 6, 15, 'Folkeskole + SFO'), ('16-24 år (ungdom)', 16, 24, 'Ungdomsuddannelse/jobcenter'), ('25-39 år (yngre voksne)', 25, 39, 'Beskæftigelse/bolig'), ('40-64 år (midaldrende)', 40, 64, 'Arbejdsmarked/sundhed'), ('65-79 år (ældre)', 65, 79, 'Ældrepleje/hjemmepleje'), ('80+ år (ældste)', 80, 125, 'Plejehjem/sygepleje')]
for col, h in enumerate(['Aldersgruppe', f'{periods[0][:4]}M{int(periods[0][5:]):02d}', f'{periods[-1][:4]}M{int(periods[-1][5:]):02d}', 'Ændring', 'Ændring %', 'Budgetområde'], 1):
    c = ws2.cell(row=3, column=col, value=h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin_border
for i, (gn, lo, hi, ba) in enumerate(bg):
    row = 4 + i
    jv = sum(data.get((f"{a} år", periods[0]), 0) for a in range(lo, hi+1))
    av = sum(data.get((f"{a} år", periods[-1]), 0) for a in range(lo, hi+1))
    ch = av - jv; cp = ch / jv if jv else 0
    ws2.cell(row=row, column=1, value=gn).font = Font(name='DM Sans', size=10, bold=True, color=TEXT); ws2.cell(row=row, column=1).border = thin_border
    for col, v in [(2, jv), (3, av)]:
        c = ws2.cell(row=row, column=col, value=v); c.font = data_font; c.alignment = Alignment(horizontal='right'); c.border = thin_border; c.number_format = '#,##0'
    c = ws2.cell(row=row, column=4, value=ch); c.font = Font(name='DM Sans', size=10, color='C0392B' if ch < 0 else GREEN_MID, bold=True); c.alignment = Alignment(horizontal='right'); c.border = thin_border; c.number_format = '+#,##0;-#,##0;0'
    c = ws2.cell(row=row, column=5, value=cp); c.font = Font(name='DM Sans', size=10, color='C0392B' if ch < 0 else GREEN_MID, bold=True); c.alignment = Alignment(horizontal='right'); c.border = thin_border; c.number_format = '+0.0%;-0.0%;0.0%'
    ws2.cell(row=row, column=6, value=ba).font = Font(name='DM Sans', size=9, color=MUTED, italic=True); ws2.cell(row=row, column=6).border = thin_border
    if i % 2 == 1:
        for col in range(1, 7): ws2.cell(row=row, column=col).fill = alt_fill
ws2.column_dimensions['A'].width = 22; ws2.column_dimensions['B'].width = 12; ws2.column_dimensions['C'].width = 12; ws2.column_dimensions['D'].width = 12; ws2.column_dimensions['E'].width = 12; ws2.column_dimensions['F'].width = 42; ws2.freeze_panes = 'A4'

# ===== SHEET 3: Vejledning =====
ws3 = wb.create_sheet('Opdatering')
ws3.column_dimensions['A'].width = 50; ws3.column_dimensions['B'].width = 80
ws3['A1'] = 'Opdatering af data'; ws3['A1'].font = Font(name='DM Serif Display', size=18, color=GREEN)
info = [
    ('', ''), ('1. Power Query (anbefalet)', ''), ('', ''),
    ('Trin 1:', 'Data > Hent data > Fra internettet'),
    ('Trin 2:', 'Indsæt URL: https://api.statbank.dk/v1/data/FOLK1AM/CSV?OMR%C3%85DE=167&K%C3%98N=TOT&ALDER=*&TID=*&lang=da'),
    ('Trin 3:', 'Filtrér "Alder i alt" fra i Power Query'),
    ('Trin 4:', 'Indlæs — fremtidig opdatering: Ctrl+Alt+F5'),
    ('', ''), ('2. VBA makro (DST_Refresh.bas)', ''), ('', ''),
    ('Importér:', 'Udvikler > Visual Basic > File > Import File > DST_Refresh.bas'),
    ('Kør:', 'Alt+F8 > RefreshDSTData'),
    ('Genvej:', 'Tilføj formular-knap og tildel RefreshDSTData'),
    ('', ''), ('3. Download CSV manuelt', ''), ('', ''),
    ('URL:', 'https://api.statbank.dk/v1/data/FOLK1AM/CSV?OMR%C3%85DE=167&K%C3%98N=TOT&ALDER=*&TID=*&lang=da'),
    ('', ''), ('API:', 'api.statbank.dk · Tabel: FOLK1AM · OMRÅDE=167 (Hvidovre) · KØN=TOT · ALDER=* · TID=*'),
]
for i, (lbl, val) in enumerate(info):
    r = 3 + i
    if lbl and not val:
        ws3.cell(row=r, column=1, value=lbl).font = Font(name='DM Sans', size=11, bold=True, color=GREEN)
    elif lbl.startswith('Trin') or lbl in ('Importér:', 'Kør:', 'Genvej:', 'URL:', 'API:'):
        ws3.cell(row=r, column=1, value=lbl).font = Font(name='DM Mono', size=10, bold=True, color=MID)
    if val:
        ws3.cell(row=r, column=2, value=val).font = Font(name='DM Sans', size=10, color=TEXT)

out = "D:/hermes-usb-portable-main/data/skills/Danmarks Statistik API/hvidovre_folk1am.xlsx"
wb.save(out)
print(f"OK: {out}")
print(f"Ark: {wb.sheetnames}")
print(f"Rådata rækker: {len(rows)}")
print(f"Aldre: {min(int(a.split()[0]) for a in ages)}–{max(int(a.split()[0]) for a in ages)}")
print(f"Perioder: {len(periods)} ({periods[0]}–{periods[-1]})")
print(f"Total {periods[0]}: {tj:,}")
print(f"Total {periods[-1]}: {ta:,}")
print(f"Ændring: {tc:+d} ({tp*100:+.1f}%)")
